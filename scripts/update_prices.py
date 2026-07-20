#!/usr/bin/env python3
"""Aggiorna prices.json con i prezzi benzina (Euro-Super 95, EUR/litro).
Fonti gratuite:
 - Italia: open data MIMIT (giornaliero) -> media nazionale benzina self
 - Paesi UE: Bollettino Petrolifero settimanale della Commissione Europea
 - Norvegia e Svizzera: nessuna fonte gratuita -> restano gli ultimi valori (correggili nel sito)
Se una fonte non risponde, i valori precedenti vengono mantenuti: il sito non resta mai senza dati.
"""
import csv, datetime, io, json, os, sys
import requests

PATH = os.path.join(os.path.dirname(__file__), "..", "prices.json")

def load():
    try:
        with open(PATH) as f: return json.load(f)
    except Exception:
        return {"prices": {}}

data = load()
prices = data.get("prices", {})
sources = {}

def set_price(code, val, src):
    try: val = float(val)
    except Exception: return
    if 0.5 < val < 4.0:
        prices[code] = round(val, 3)
        sources[code] = src
        print(f"  {code}: {val:.3f} EUR/L  ({src})")

# ---------- ITALIA: MIMIT, prezzi comunicati alle 8 di oggi ----------
print("Italia (MIMIT)…")
try:
    r = requests.get("https://www.mimit.gov.it/images/exportCSV/prezzo_alle_8.csv",
                     timeout=90, headers={"User-Agent": "caponord-prezzi"})
    r.raise_for_status()
    tot = n = 0.0
    for row in csv.reader(io.StringIO(r.text), delimiter=";"):
        # idImpianto;descCarburante;prezzo;isSelf;dtComu
        if len(row) >= 4 and row[1].strip().lower() == "benzina" and row[3].strip() == "1":
            try:
                v = float(row[2].replace(",", "."))
                if 1.0 < v < 3.0:
                    tot += v; n += 1
            except ValueError:
                pass
    if n > 1000:
        set_price("IT", tot / n, f"MIMIT media su {int(n)} impianti")
    else:
        print("  IT: troppi pochi dati, mantengo il valore precedente")
except Exception as e:
    print("  IT: fonte non raggiungibile:", e)

# ---------- PAESI UE: Bollettino Petrolifero settimanale ----------
# Se il link cambia (succede ogni tanto), aggiorna questa lista:
EU_URLS = [
    "https://ec.europa.eu/energy/observatory/reports/latest_prices_with_taxes.xlsx",
    "https://energy.ec.europa.eu/system/files/latest_prices_with_taxes.xlsx",
]
EU_WANTED = {"DE", "DK", "SE", "FI", "EE", "LV", "LT", "PL"}
print("Bollettino UE…")
for url in EU_URLS:
    try:
        r = requests.get(url, timeout=90, headers={"User-Agent": "caponord-prezzi"})
        r.raise_for_status()
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content), data_only=True)
        found = set()
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if not row or not isinstance(row[0], str):
                    continue
                cc = row[0].strip().upper()[:2]
                if cc in EU_WANTED and cc not in found:
                    # la benzina 95 è in EUR per 1000 litri: primo numero plausibile della riga
                    for cell in row[1:]:
                        if isinstance(cell, (int, float)) and 800 < cell < 4000:
                            set_price(cc, cell / 1000, "Bollettino UE settimanale")
                            found.add(cc)
                            break
        if found:
            break
    except Exception as e:
        print("  UE: tentativo fallito:", e)
else:
    print("  UE: nessun link ha risposto, mantengo i valori precedenti")

# ---------- salva ----------
out = {
    "updated": datetime.date.today().isoformat(),
    "prices": prices,
    "sources": sources or data.get("sources", {}),
    "note": "NO e CH senza fonte automatica gratuita: correggili a mano nel sito quando serve.",
}
with open(PATH, "w") as f:
    json.dump(out, f, indent=1, ensure_ascii=False)
print("Salvato prices.json:", out["updated"])
